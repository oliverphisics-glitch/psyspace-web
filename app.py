from flask import Flask, render_template, request, redirect, url_for, session, send_file, flash
import sqlite3
import hashlib
import os
import math
from datetime import datetime

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, RadarChart, Reference
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_OK = True
except Exception:
    OPENPYXL_OK = False


app = Flask(__name__)
app.secret_key = "psyspace_web_2026"

DB = "psyspace_web.db"
SENHA_PESQUISADOR = "psyspace2026"

DIMENSOES = [
    "Atenção",
    "Flexibilidade",
    "Autorregulação",
    "Persistência",
    "Cooperação",
    "Planejamento"
]

NOME_PROJETO = (
    "Simulador Educacional Inspirado em Simulações Aeroespaciais "
    "para Avaliação de Carga Cognitiva e Funções Executivas em Estudantes"
)

ESCALA = [
    ("Atenção", "Consigo manter a concentração durante atividades longas.", 0),
    ("Atenção", "Evito distrações quando estou realizando uma tarefa importante.", 0),
    ("Atenção", "Consigo concluir atividades sem perder o foco.", 0),
    ("Atenção", "Mantenho a atenção mesmo em ambientes movimentados.", 0),
    ("Atenção", "Retomo rapidamente uma atividade após uma interrupção.", 0),
    ("Atenção", "Frequentemente me distraio durante tarefas importantes.", 1),
    ("Atenção", "Tenho dificuldade em acompanhar explicações longas.", 1),
    ("Atenção", "Consigo prestar atenção aos detalhes.", 0),
    ("Atenção", "Perco o foco facilmente quando algo chama minha atenção.", 1),
    ("Atenção", "Consigo acompanhar instruções sem me perder.", 0),

    ("Flexibilidade", "Adapto-me facilmente a situações novas.", 0),
    ("Flexibilidade", "Consigo encontrar diferentes soluções para um problema.", 0),
    ("Flexibilidade", "Mudo de estratégia quando percebo que algo não funciona.", 0),
    ("Flexibilidade", "Gosto de aprender novas formas de realizar tarefas.", 0),
    ("Flexibilidade", "Aceito mudanças de planos sem grandes dificuldades.", 0),
    ("Flexibilidade", "Fico incomodado quando preciso mudar minha rotina.", 1),
    ("Flexibilidade", "Tenho dificuldade em aceitar novas ideias.", 1),
    ("Flexibilidade", "Consigo aprender com meus erros.", 0),
    ("Flexibilidade", "Costumo considerar diferentes pontos de vista.", 0),
    ("Flexibilidade", "Adapto-me rapidamente a mudanças inesperadas.", 0),

    ("Autorregulação", "Consigo controlar minhas emoções em situações difíceis.", 0),
    ("Autorregulação", "Mantenho a calma diante de problemas.", 0),
    ("Autorregulação", "Lido bem com críticas construtivas.", 0),
    ("Autorregulação", "Consigo controlar minha impulsividade.", 0),
    ("Autorregulação", "Penso antes de agir.", 0),
    ("Autorregulação", "Costumo reagir sem pensar.", 1),
    ("Autorregulação", "Perco o controle emocional com facilidade.", 1),
    ("Autorregulação", "Consigo me recuperar rapidamente após uma frustração.", 0),
    ("Autorregulação", "Separo emoções de decisões importantes.", 0),
    ("Autorregulação", "Mantenho o equilíbrio emocional em situações desafiadoras.", 0),

    ("Persistência", "Continuo tentando mesmo quando encontro dificuldades.", 0),
    ("Persistência", "Não desisto facilmente dos meus objetivos.", 0),
    ("Persistência", "Aprendo com os erros e continuo tentando.", 0),
    ("Persistência", "Mantenho meus esforços mesmo diante de fracassos.", 0),
    ("Persistência", "Tenho determinação para concluir tarefas.", 0),
    ("Persistência", "Costumo abandonar tarefas difíceis.", 1),
    ("Persistência", "Desanimo rapidamente quando erro.", 1),
    ("Persistência", "Encaro desafios como oportunidades de aprendizagem.", 0),
    ("Persistência", "Busco alternativas quando algo não dá certo.", 0),
    ("Persistência", "Persisto até concluir aquilo que comecei.", 0),

    ("Cooperação", "Trabalho bem em equipe.", 0),
    ("Cooperação", "Escuto opiniões diferentes das minhas.", 0),
    ("Cooperação", "Respeito as ideias dos colegas.", 0),
    ("Cooperação", "Consigo colaborar para alcançar objetivos comuns.", 0),
    ("Cooperação", "Comunico minhas ideias com clareza.", 0),
    ("Cooperação", "Evito trabalhar em grupo.", 1),
    ("Cooperação", "Tenho dificuldade em ouvir opiniões diferentes.", 1),
    ("Cooperação", "Ajudo colegas quando necessário.", 0),
    ("Cooperação", "Participo ativamente de atividades coletivas.", 0),
    ("Cooperação", "Busco resolver conflitos de forma respeitosa.", 0),

    ("Planejamento", "Organizo meu tempo para cumprir tarefas.", 0),
    ("Planejamento", "Costumo planejar minhas atividades.", 0),
    ("Planejamento", "Defino metas para alcançar objetivos.", 0),
    ("Planejamento", "Consigo priorizar tarefas importantes.", 0),
    ("Planejamento", "Cumpro prazos estabelecidos.", 0),
    ("Planejamento", "Deixo tarefas para a última hora.", 1),
    ("Planejamento", "Tenho dificuldade em organizar minhas atividades.", 1),
    ("Planejamento", "Acompanho meu próprio progresso.", 0),
    ("Planejamento", "Assumo responsabilidade pelas minhas decisões.", 0),
    ("Planejamento", "Planejo etapas antes de iniciar uma tarefa.", 0),
]


RECOMENDACOES = {
    "Atenção": [
        "Dividir tarefas longas em etapas curtas.",
        "Usar instruções objetivas e visuais.",
        "Inserir pausas cognitivas durante atividades extensas.",
        "Reduzir distrações no ambiente de aprendizagem."
    ],
    "Flexibilidade": [
        "Propor problemas com mais de uma solução possível.",
        "Discutir diferentes estratégias de resolução.",
        "Estimular revisão de hipóteses após erros.",
        "Criar situações com mudanças graduais de regras."
    ],
    "Autorregulação": [
        "Trabalhar metas de curto prazo.",
        "Usar feedback frequente e objetivo.",
        "Estimular autoavaliação após cada atividade.",
        "Criar momentos de reflexão sobre erros."
    ],
    "Persistência": [
        "Usar desafios graduais com dificuldade crescente.",
        "Valorizar o processo e não apenas o resultado.",
        "Oferecer reforço positivo diante de tentativas.",
        "Trabalhar aprendizagem pelo erro."
    ],
    "Cooperação": [
        "Utilizar aprendizagem por pares.",
        "Organizar grupos pequenos com papéis definidos.",
        "Trabalhar escuta ativa e comunicação científica.",
        "Promover resolução coletiva de problemas."
    ],
    "Planejamento": [
        "Usar checklists de tarefas.",
        "Trabalhar cronogramas visuais.",
        "Definir etapas claras antes da atividade.",
        "Ensinar organização do tempo e priorização."
    ]
}


def conectar():
    return sqlite3.connect(DB)


def hash_telefone(telefone):
    telefone_limpo = "".join(filter(str.isdigit, telefone))
    return hashlib.sha256(telefone_limpo.encode()).hexdigest()


def criar_banco():
    conn = conectar()
    cur = conn.cursor()

    cur.execute("""
    CREATE TABLE IF NOT EXISTS participantes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        telefone_hash TEXT UNIQUE,
        idade INTEGER,
        sexo TEXT,
        serie TEXT,
        etnia TEXT,
        clube TEXT,
        pesquisa TEXT,
        consentimento TEXT,
        pre_concluido INTEGER DEFAULT 0,
        pos_concluido INTEGER DEFAULT 0,
        data_cadastro TEXT
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS itens (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        dimensao TEXT,
        texto TEXT,
        invertida INTEGER
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS respostas_pre (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        telefone_hash TEXT,
        id_item INTEGER,
        resposta INTEGER,
        resposta_corrigida INTEGER,
        data_resposta TEXT
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS respostas_pos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        telefone_hash TEXT,
        id_item INTEGER,
        resposta INTEGER,
        resposta_corrigida INTEGER,
        data_resposta TEXT
    )
    """)

    cur.execute("SELECT COUNT(*) FROM itens")
    total = cur.fetchone()[0]
    if total == 0:
        for dimensao, texto, invertida in ESCALA:
            cur.execute(
                "INSERT INTO itens (dimensao, texto, invertida) VALUES (?, ?, ?)",
                (dimensao, texto, invertida)
            )

    conn.commit()
    conn.close()


def buscar_participante(hash_tel):
    conn = conectar()
    cur = conn.cursor()
    cur.execute("SELECT * FROM participantes WHERE telefone_hash=?", (hash_tel,))
    p = cur.fetchone()
    conn.close()
    return p


def cadastrar_participante(hash_tel, form):
    conn = conectar()
    cur = conn.cursor()
    cur.execute("""
    INSERT INTO participantes (
        telefone_hash, idade, sexo, serie, etnia, clube, pesquisa, consentimento, data_cadastro
    )
    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        hash_tel,
        form.get("idade"),
        form.get("sexo"),
        form.get("serie"),
        form.get("etnia"),
        form.get("clube"),
        form.get("pesquisa"),
        "Sim",
        datetime.now().strftime("%d/%m/%Y %H:%M")
    ))
    conn.commit()
    conn.close()


def buscar_item(indice):
    conn = conectar()
    cur = conn.cursor()
    cur.execute("SELECT id, dimensao, texto, invertida FROM itens ORDER BY id LIMIT 1 OFFSET ?", (indice,))
    item = cur.fetchone()
    conn.close()
    return item


def total_itens():
    conn = conectar()
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM itens")
    total = cur.fetchone()[0]
    conn.close()
    return total


def corrigir(resposta, invertida):
    return 6 - resposta if invertida == 1 else resposta


def salvar_resposta(tabela, hash_tel, id_item, resposta, corrigida):
    conn = conectar()
    cur = conn.cursor()
    cur.execute(f"DELETE FROM {tabela} WHERE telefone_hash=? AND id_item=?", (hash_tel, id_item))
    cur.execute(f"""
    INSERT INTO {tabela} (telefone_hash, id_item, resposta, resposta_corrigida, data_resposta)
    VALUES (?, ?, ?, ?, ?)
    """, (hash_tel, id_item, resposta, corrigida, datetime.now().strftime("%d/%m/%Y %H:%M")))
    conn.commit()
    conn.close()


def marcar_concluido(hash_tel, campo):
    conn = conectar()
    cur = conn.cursor()
    cur.execute(f"UPDATE participantes SET {campo}=1 WHERE telefone_hash=?", (hash_tel,))
    conn.commit()
    conn.close()


def calcular_resultados(hash_tel, tabela):
    conn = conectar()
    cur = conn.cursor()
    cur.execute(f"""
    SELECT i.dimensao, r.resposta_corrigida
    FROM {tabela} r
    JOIN itens i ON i.id = r.id_item
    WHERE r.telefone_hash=?
    """, (hash_tel,))
    rows = cur.fetchall()
    conn.close()

    dados = {}
    for d, r in rows:
        dados.setdefault(d, []).append(r)

    finais = {}
    for d, vals in dados.items():
        finais[d] = round((sum(vals) / (len(vals) * 5)) * 100, 1)
    return finais


def participantes_pre_pos():
    conn = conectar()
    cur = conn.cursor()
    cur.execute("""
    SELECT telefone_hash, idade, sexo, serie, etnia, clube, pesquisa, data_cadastro
    FROM participantes
    WHERE pre_concluido=1 AND pos_concluido=1
    ORDER BY id
    """)
    rows = cur.fetchall()
    conn.close()
    return rows


def contagens():
    conn = conectar()
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM participantes")
    cadastrados = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM participantes WHERE pre_concluido=1")
    pre = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM participantes WHERE pre_concluido=1 AND pos_concluido=1")
    prepos = cur.fetchone()[0]
    conn.close()
    return cadastrados, pre, prepos


def dados_coletivos():
    rows = participantes_pre_pos()
    longo = []
    largo = []

    for idx, p in enumerate(rows, start=1):
        hash_tel, idade, sexo, serie, etnia, clube, pesquisa, data = p
        pre = calcular_resultados(hash_tel, "respostas_pre")
        pos = calcular_resultados(hash_tel, "respostas_pos")

        linha = {
            "participante": f"P{idx:03d}",
            "idade": idade,
            "sexo": sexo,
            "serie": serie,
            "etnia": etnia,
            "clube": clube,
            "pesquisa": pesquisa,
            "data_cadastro": data
        }

        for d in DIMENSOES:
            vp = pre.get(d, 0)
            vs = pos.get(d, 0)
            delta = round(vs - vp, 1)
            longo.append({"participante": f"P{idx:03d}", "dimensao": d, "pre": vp, "pos": vs, "variacao": delta})
            linha[f"{d}_pre"] = vp
            linha[f"{d}_pos"] = vs
            linha[f"{d}_variacao"] = delta
        largo.append(linha)

    return longo, largo


def medias_coletivas():
    longo, _ = dados_coletivos()
    saida = {}
    for d in DIMENSOES:
        pre = [l["pre"] for l in longo if l["dimensao"] == d]
        pos = [l["pos"] for l in longo if l["dimensao"] == d]
        if pre and pos:
            saida[d] = {
                "pre": round(sum(pre) / len(pre), 1),
                "pos": round(sum(pos) / len(pos), 1),
                "ganho": round((sum(pos) / len(pos)) - (sum(pre) / len(pre)), 1)
            }
    return saida


def interpretar_coletivo():
    med = medias_coletivas()
    if not med:
        return {
            "titulo": "Coleta em andamento",
            "texto": "Ainda não há participantes com pré e pós concluídos.",
            "recomendacoes": []
        }

    pos = {d: med[d]["pos"] for d in med}
    menor = min(pos.items(), key=lambda x: x[1])
    maior = max(pos.items(), key=lambda x: x[1])

    recs = RECOMENDACOES.get(menor[0], [])

    return {
        "titulo": f"Foco pedagógico prioritário: {menor[0]}",
        "texto": f"A dimensão mais forte no pós-teste foi {maior[0]} ({maior[1]}%). A dimensão que exige maior atenção pedagógica foi {menor[0]} ({menor[1]}%).",
        "recomendacoes": recs
    }


def exportar_excel():
    if not OPENPYXL_OK:
        return None

    longo, largo = dados_coletivos()
    med = medias_coletivas()
    interp = interpretar_coletivo()
    cadastrados, pre, prepos = contagens()

    caminho = "psyspace_web_resultados.xlsx"
    wb = Workbook()

    ws = wb.active
    ws.title = "Resumo"
    ws["A1"] = "PSYSPACE"
    ws["A2"] = NOME_PROJETO
    ws["A4"] = "Participantes cadastrados"
    ws["B4"] = cadastrados
    ws["A5"] = "Pré concluído"
    ws["B5"] = pre
    ws["A6"] = "Pré e Pós concluídos"
    ws["B6"] = prepos

    ws2 = wb.create_sheet("Resultados Individuais")
    headers = ["participante", "idade", "sexo", "serie", "etnia", "clube", "pesquisa", "data_cadastro"]
    for d in DIMENSOES:
        headers += [f"{d}_pre", f"{d}_pos", f"{d}_variacao"]
    ws2.append(headers)
    for l in largo:
        ws2.append([l.get(h, "") for h in headers])

    ws3 = wb.create_sheet("Base Longa")
    ws3.append(["participante", "dimensao", "pre", "pos", "variacao"])
    for l in longo:
        ws3.append([l["participante"], l["dimensao"], l["pre"], l["pos"], l["variacao"]])

    ws4 = wb.create_sheet("Estatísticas")
    ws4.append(["dimensao", "media_pre", "media_pos", "ganho"])
    for d in DIMENSOES:
        v = med.get(d, {"pre": 0, "pos": 0, "ganho": 0})
        ws4.append([d, v["pre"], v["pos"], v["ganho"]])

    ws5 = wb.create_sheet("Gráficos")
    ws5.append(["dimensao", "pre", "pos", "ganho"])
    for d in DIMENSOES:
        v = med.get(d, {"pre": 0, "pos": 0, "ganho": 0})
        ws5.append([d, v["pre"], v["pos"], v["ganho"]])

    bar = BarChart()
    bar.title = "Comparação coletiva Pré x Pós"
    data = Reference(ws5, min_col=2, max_col=3, min_row=1, max_row=7)
    cats = Reference(ws5, min_col=1, min_row=2, max_row=7)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    bar.height = 8
    bar.width = 18
    ws5.add_chart(bar, "F2")

    radar = RadarChart()
    radar.title = "Radar coletivo Pré x Pós"
    data_radar = Reference(ws5, min_col=2, max_col=3, min_row=1, max_row=7)
    radar.add_data(data_radar, titles_from_data=True)
    radar.set_categories(cats)
    radar.height = 8
    radar.width = 18
    ws5.add_chart(radar, "F18")

    ws6 = wb.create_sheet("Interpretação Pedagógica")
    ws6["A1"] = "Interpretação Pedagógica Automatizada"
    ws6["A3"] = interp["titulo"]
    ws6["A4"] = interp["texto"]
    ws6["A6"] = "Recomendações"
    linha = 7
    for rec in interp["recomendacoes"]:
        ws6[f"A{linha}"] = f"• {rec}"
        linha += 1

    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                max_len = max(max_len, len(str(cell.value)) if cell.value is not None else 0)
            sheet.column_dimensions[col_letter].width = min(max_len + 3, 35)

    wb.save(caminho)
    return caminho


def gerar_grafico(tipo):
    med = medias_coletivas()
    os.makedirs("static/graficos", exist_ok=True)
    caminho = f"static/graficos/{tipo}.png"

    plt.figure(figsize=(9, 5))

    if not med:
        plt.text(0.5, 0.5, "Coleta em andamento\nSem dados coletivos suficientes", ha="center", va="center", fontsize=14)
        plt.axis("off")
    else:
        dims = DIMENSOES
        pre = [med.get(d, {}).get("pre", 0) for d in dims]
        pos = [med.get(d, {}).get("pos", 0) for d in dims]
        ganho = [med.get(d, {}).get("ganho", 0) for d in dims]

        if tipo == "barras":
            x = range(len(dims))
            plt.bar([i - 0.18 for i in x], pre, width=0.36, label="Pré")
            plt.bar([i + 0.18 for i in x], pos, width=0.36, label="Pós")
            plt.xticks(list(x), dims, rotation=20)
            plt.ylim(0, 100)
            plt.ylabel("Média (%)")
            plt.title("Comparação coletiva Pré × Pós")
            plt.legend()
        elif tipo == "ganhos":
            plt.bar(dims, ganho)
            plt.axhline(0, color="black", linewidth=0.8)
            plt.xticks(rotation=20)
            plt.ylabel("Variação média (%)")
            plt.title("Ganho médio por dimensão")
        elif tipo == "radar":
            plt.close()
            fig = plt.figure(figsize=(7, 5))
            ax = fig.add_subplot(111, polar=True)
            angles = [n / float(len(dims)) * 2 * math.pi for n in range(len(dims))]
            angles += angles[:1]
            pre_r = pre + pre[:1]
            pos_r = pos + pos[:1]
            ax.plot(angles, pre_r, linewidth=2, label="Pré")
            ax.fill(angles, pre_r, alpha=0.15)
            ax.plot(angles, pos_r, linewidth=2, label="Pós")
            ax.fill(angles, pos_r, alpha=0.15)
            ax.set_xticks(angles[:-1])
            ax.set_xticklabels(dims)
            ax.set_ylim(0, 100)
            ax.set_title("Radar coletivo Pré × Pós")
            ax.legend(loc="upper right")

    plt.tight_layout()
    plt.savefig(caminho, dpi=150)
    plt.close()
    return caminho


@app.context_processor
def inject_globals():
    return dict(nome_projeto=NOME_PROJETO)


@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        telefone = request.form.get("telefone", "")
        telefone_limpo = "".join(filter(str.isdigit, telefone))
        if len(telefone_limpo) < 8:
            flash("Digite um telefone válido.")
            return redirect(url_for("index"))

        hash_tel = hash_telefone(telefone_limpo)
        session["telefone_hash"] = hash_tel

        p = buscar_participante(hash_tel)
        if p is None:
            return redirect(url_for("cadastro"))

        pre = p[9]
        pos = p[10]

        if pre == 0:
            session["modo"] = "pre"
            session["indice"] = 0
            return redirect(url_for("questionario"))
        if pos == 0:
            session["modo"] = "pos"
            session["indice"] = 0
            return redirect(url_for("questionario"))

        return redirect(url_for("finalizado"))

    return render_template("index.html")


@app.route("/cadastro", methods=["GET", "POST"])
def cadastro():
    if "telefone_hash" not in session:
        return redirect(url_for("index"))

    if request.method == "POST":
        if request.form.get("consentimento") != "on":
            flash("É necessário aceitar o termo de consentimento.")
            return redirect(url_for("cadastro"))

        cadastrar_participante(session["telefone_hash"], request.form)
        session["modo"] = "pre"
        session["indice"] = 0
        return redirect(url_for("questionario"))

    return render_template("cadastro.html")


@app.route("/questionario", methods=["GET", "POST"])
def questionario():
    if "telefone_hash" not in session:
        return redirect(url_for("index"))

    modo = session.get("modo", "pre")
    indice = int(session.get("indice", 0))
    total = total_itens()

    if request.method == "POST":
        resposta = int(request.form.get("resposta", 0))
        item = buscar_item(indice)
        if item and resposta:
            id_item, dimensao, texto, invertida = item
            tabela = "respostas_pre" if modo == "pre" else "respostas_pos"
            salvar_resposta(tabela, session["telefone_hash"], id_item, resposta, corrigir(resposta, invertida))

        acao = request.form.get("acao")
        if acao == "anterior":
            indice = max(0, indice - 1)
        else:
            indice += 1

        session["indice"] = indice
        return redirect(url_for("questionario"))

    if indice >= total:
        if modo == "pre":
            marcar_concluido(session["telefone_hash"], "pre_concluido")
            return redirect(url_for("resultado", modo="pre"))
        else:
            marcar_concluido(session["telefone_hash"], "pos_concluido")
            return redirect(url_for("resultado", modo="pos"))

    item = buscar_item(indice)
    progresso = int(((indice + 1) / total) * 100)

    return render_template("questionario.html", item=item, indice=indice, total=total, progresso=progresso, modo=modo)


@app.route("/resultado/<modo>")
def resultado(modo):
    if "telefone_hash" not in session:
        return redirect(url_for("index"))
    tabela = "respostas_pre" if modo == "pre" else "respostas_pos"
    resultados = calcular_resultados(session["telefone_hash"], tabela)
    return render_template("resultado.html", resultados=resultados, modo=modo)


@app.route("/finalizado")
def finalizado():
    return render_template("finalizado.html")


@app.route("/pesquisador", methods=["GET", "POST"])
def pesquisador_login():
    if request.method == "POST":
        senha = request.form.get("senha", "")
        if senha == SENHA_PESQUISADOR:
            session["pesquisador"] = True
            return redirect(url_for("painel"))
        flash("Senha incorreta.")
    return render_template("pesquisador_login.html")


@app.route("/painel")
def painel():
    if not session.get("pesquisador"):
        return redirect(url_for("pesquisador_login"))

    cadastrados, pre, prepos = contagens()
    med = medias_coletivas()
    interp = interpretar_coletivo()

    grafico = request.args.get("grafico", "barras")
    if grafico not in ["barras", "radar", "ganhos"]:
        grafico = "barras"
    caminho_grafico = gerar_grafico(grafico)

    return render_template(
        "painel.html",
        cadastrados=cadastrados,
        pre=pre,
        prepos=prepos,
        med=med,
        interp=interp,
        grafico=grafico,
        caminho_grafico="/" + caminho_grafico
    )


@app.route("/exportar_excel")
def exportar_excel_rota():
    if not session.get("pesquisador"):
        return redirect(url_for("pesquisador_login"))

    caminho = exportar_excel()
    if not caminho:
        flash("Instale openpyxl: pip install openpyxl")
        return redirect(url_for("painel"))

    return send_file(caminho, as_attachment=True)


@app.route("/logout_pesquisador")
def logout_pesquisador():
    session.pop("pesquisador", None)
    return redirect(url_for("index"))


if __name__ == "__main__":
    criar_banco()
    app.run(host="0.0.0.0", port=5000, debug=True)
